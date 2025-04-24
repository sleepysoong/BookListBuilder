#!/usr/bin/env python3
import concurrent.futures
import sys
import requests
import xlsxwriter
import openpyxl
import os
from io import BytesIO
from dataclasses import dataclass, field
from PIL import Image, ImageFont
import json
from typing import Optional, List, Tuple, Dict, Any, Callable
from enum import Enum, auto
from urllib.parse import urlparse, parse_qs
import yaml
from datetime import datetime
from xlsxwriter.format import Format

VERSION: str = "1.4.4"


class LibraryBookStatus(Enum):
    EXISTS = auto()
    NOT_EXISTS = auto()
    UNKNOWN = auto()

@dataclass
class Book:
    item_id: Optional[int] = None # 알라딘 아이템 아이디
    order: int = 0
    key: Optional[int] = None # 도서관 도서 키
    species_key: Optional[int] = None # 도서관 도서 종류 키
    title: str = ''
    cover: Optional[BytesIO] = None
    author: str = ''
    publisher: str = ''
    isbn13: str = ''
    standard_price: int = 0
    publish_date: str = ''
    description: str = ''
    rating_score: float = 0.0
    rating_count: int = 0
    sales_point: int = 0
    category: str = ''
    sheet_name: str = ''
    library_status: LibraryBookStatus = LibraryBookStatus.UNKNOWN
    memo: str = ''

class Column:
    def __init__(self, header: str, getter: Callable[[Book], Any], fmt_key: Optional[str]):
        self.header: str = header
        self.getter: Callable[[Book], Any] = getter
        self.fmt_key: Optional[str] = fmt_key

class FormatManager:
    def __init__(self, workbook: xlsxwriter.Workbook, font_size_pt: int):
        base_args: Dict[str, Any] = {'font_size': font_size_pt, 'text_wrap': True}
        self.fmts: Dict[str, Format] = {
            'header': workbook.add_format({**base_args, 'bold': True, 'bg_color': '#D3D3D3', 'align': 'center', 'valign': 'vcenter'}),
            'center': workbook.add_format({**base_args, 'align': 'center', 'valign': 'vcenter'}),
            'left': workbook.add_format({**base_args, 'align': 'left', 'valign': 'vcenter'}),
            'price': workbook.add_format({**base_args, 'num_format': '#,##0"원"', 'align': 'center', 'valign': 'vcenter'}),
            'hyperlink': workbook.add_format({**base_args, 'font_color': 'blue', 'underline': 1, 'align': 'center', 'valign': 'vcenter'}),
            'sales_point': workbook.add_format({**base_args, 'num_format': '#,##0', 'align': 'center', 'valign': 'vcenter'}),
        }
        # 도서관 소장 도서 행 강조용 노란색 배경 포맷 추가
        self.fmts['highlight'] = workbook.add_format({**base_args, 'bg_color': '#FFFF00'})
        # 5년 이전 도서 연두색 배경 포맷 추가
        self.fmts['oldbook'] = workbook.add_format({**base_args, 'bg_color': '#CCFFCC'})

    def get(self, key: str) -> Optional[Format]:
        return self.fmts.get(key)
    
COLUMNS: List[Column] = [
    Column('', lambda b: b.cover, None),
    Column('도서', lambda b: b.title, 'center'),
    Column('저자', lambda b: b.author, 'center'),
    Column('출판사', lambda b: b.publisher, 'center'),
    Column('ISBN13', lambda b: b.isbn13, 'center'),
    Column('정가', lambda b: b.standard_price, 'price'),
    Column('출판일', lambda b: b.publish_date, 'center'),
    Column('설명', lambda b: b.description, 'left'),
    Column('평점', lambda b: f'{b.rating_score:.1f} {"★"*round(b.rating_score/2)+"☆"*(5-round(b.rating_score/2))} ({b.rating_count})', 'center'),
    Column('판매 지수', lambda b: b.sales_point, 'sales_point'),
    Column('카테고리', lambda b: b.category, 'left'),
    Column('교내 도서관 소장', lambda b: 'O' if b.library_status == LibraryBookStatus.EXISTS else ('X' if b.library_status == LibraryBookStatus.NOT_EXISTS else '?'), 'center'),
    Column('메모', lambda b: b.memo, 'left'),
]

def update_library_status(book: Book, neis_code: str, prov_code: str, session: requests.Session, timeout: int = 10) -> None:
    url: str = "https://read365.edunet.net/alpasq/api/search"

    payload: Dict[str, Any] = {
        "searchKeyword": book.isbn13,
        "neisCode": [neis_code],
        "provCode": prov_code,
        "coverYn": "N"
    }

    headers: Dict[str, str] = {"Content-Type": "application/json"}

    try:
        response: requests.Response = session.post(url, json=payload, headers=headers, timeout=timeout)
        response.raise_for_status()

        data: Dict[str, Any] = response.json()

        results: List[Dict[str, Any]] = data.get("data", {}).get("bookList", [])

        for item in results:
            if item.get("isbn") == book.isbn13:
                book.library_status = LibraryBookStatus.EXISTS
                book.key = item.get('bookKey')
                book.species_key = item.get('speciesKey')
                print(f"> [조회 성공][도서관] ISBN {book.isbn13}: 소장하고 있는 도서에요 ㅡ bookKey: {book.key}, speciesKey: {book.species_key}")
                return
            
        book.library_status = LibraryBookStatus.NOT_EXISTS
        print(f"> [조회 성공][도서관] ISBN {book.isbn13}: 소장하고 있지 않은 도서에요.")
        return
    
    except requests.exceptions.HTTPError as http_err:
        print(f"> [조회 실패][도서관] ISBN {book.isbn13}: HTTP 오류가 발생했어요 - {http_err.response.status_code} {http_err.response.reason}")
        book.library_status = LibraryBookStatus.UNKNOWN
        return
    
    except requests.exceptions.RequestException as e:
        print(f"> [조회 실패][도서관] ISBN {book.isbn13}: 네트워크 오류가 발생했어요 - {e}")
        book.library_status = LibraryBookStatus.UNKNOWN
        return
    
    except Exception as e:
        print(f"> [조회 실패][도서관] ISBN {book.isbn13}: 예상치 못한 오류가 발생했어요 - {e}")
        book.library_status = LibraryBookStatus.UNKNOWN
        return

def update_book_info(book: Book, aladin_api_key: str, session: requests.Session, timeout: int = 5) -> None:
    if book.item_id:
        url = (
            f"http://www.aladin.co.kr/ttb/api/ItemLookUp.aspx"
            f"?ttbkey={aladin_api_key}&itemIdType=ItemId&ItemId={book.item_id}"
            "&output=js&Version=20131101&OptResult=Story,categoryIdList,"
            "bestSellerRank,ratingInfo,reviewList"
        )
    else:
        url = (
            f"http://www.aladin.co.kr/ttb/api/ItemLookUp.aspx"
            f"?ttbkey={aladin_api_key}&itemIdType=ISBN&ItemId={book.isbn13}"
            "&output=js&Version=20131101&OptResult=Story,categoryIdList,"
            "bestSellerRank,ratingInfo,reviewList"
        )

    try:
        resp: requests.Response = session.get(url, timeout=timeout)
        resp.raise_for_status()

        items: List[Dict[str, Any]] = resp.json().get("item", [])

        if not items:
            print(f"> [조회 실패][알라딘] ISBN {book.isbn13}: 데이터를 찾을 수 없어요.")
            return
        
        item: Dict[str, Any] = items[0]

        desc: str = item.get("description", "").strip()
        rating_info: Dict[str, Any] = item.get("subInfo", {}).get("ratingInfo", {})
        score: float = float(rating_info.get("ratingScore", 0))
        count: int = int(rating_info.get("ratingCount", 0))

        book.title = item.get("title", "")
        book.item_id = item.get("itemId", "")
        book.author = item.get("author", "")
        book.publisher = item.get("publisher", "")
        book.isbn13 = item.get("isbn13", "")
        book.standard_price = int(item.get("priceStandard", 0))
        book.publish_date = item.get("pubDate", "")
        book.description = desc
        book.rating_score = score
        book.rating_count = count
        book.sales_point = item.get("salesPoint", 0)
        book.category = item.get("categoryName", "")

        print(f"> [조회 성공][알라딘] ISBN {book.isbn13}: '{book.title}' 정보를 가져왔어요.")

        if cover_url := item.get("cover"):
            try:
                cover_resp: requests.Response = session.get(cover_url, timeout=timeout)
                cover_resp.raise_for_status()

                book.cover = BytesIO(cover_resp.content)

            except requests.exceptions.RequestException as e:
                print(f"> [오류][알라딘] ISBN {book.isbn13} 커버 이미지를 가져오는 중 오류가 발생했어요: {e}")
                book.cover = None

    except requests.exceptions.RequestException as e:
        print(f"> [조회 실패][알라딘] ISBN {book.isbn13}: 정보 가져오는 중 오류가 발생했어요 - {e}")
        return
    
    except json.JSONDecodeError:
        print(f"> [조회 실패][알라딘] ISBN {book.isbn13}: 응답이 유효한 JSON 형식이 아니에요.")
        return
    
    except Exception as e:
         print(f"> [조회 실패][알라딘] ISBN {book.isbn13}: 정보 처리 중 예상치 못한 오류가 발생했어요 - {e}")
         return

def get_text_px(text: str, font: ImageFont.FreeTypeFont) -> Tuple[int, int]:
    char_width_avg: float = font.getlength('A')
    char_height: int = font.size

    max_width: int = 0

    lines: List[str] = text.split('\n')

    for line in lines:
        line_width: float = 0

        for char in line:
            line_width += char_width_avg * 2 if ord(char) > 127 else char_width_avg

        max_width = max(max_width, line_width)

    return int(max_width), int(len(lines) * char_height * 1.2)

def col_to_px(width: float) -> int:
    return int(width * 7 + 5)

def row_to_px(height: float) -> int:
    return int(height * 96 / 72)

def create(
        books: List[Book],
        output: str,
        font_size_pt: int,
        aladin_api_key: Optional[str] = None,
        neis_code: Optional[str] = None,
        prov_code: Optional[str] = None,
        school_name: Optional[str] = None
    ) -> None:
    font: ImageFont.ImageFont = ImageFont.load_default()
    workbook: xlsxwriter.Workbook = xlsxwriter.Workbook(output, {'default_date_format': 'yyyy-mm-dd'})
    fm: FormatManager = FormatManager(workbook, font_size_pt)
    session: requests.Session = requests.Session()

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures: List[concurrent.futures.Future] = []

        for book in books:
            if book.item_id is not None and book.isbn13 == "":
                # 알라딘 정보 조회 후 도서관 상태 조회
                info_future = executor.submit(update_book_info, book, aladin_api_key, session)
                concurrent.futures.wait([info_future])
                futures.append(executor.submit(update_library_status, book, neis_code, prov_code, session))
            else:
                # 알라딘과 도서관 동시에 조회
                futures.append(executor.submit(update_book_info, book, aladin_api_key, session))
                futures.append(executor.submit(update_library_status, book, neis_code, prov_code, session))

        concurrent.futures.wait(futures)

    print(f"@@@@@ 총 {len(books)}권의 책 정보를 성공적으로 가져왔어요.")

    sheet_sequence: List[str] = []
    for book in books:
        if book.sheet_name not in sheet_sequence:
            sheet_sequence.append(book.sheet_name)

    for sheet_name in sheet_sequence:
        group_books: List[Book] = [book for book in books if book.sheet_name == sheet_name]
        group_books.sort(key=lambda x: x.order)

        worksheet: xlsxwriter.Worksheet = workbook.add_worksheet(sheet_name)

        img_idx: int = 0
        img_col_width_char: int = 16

        col_widths: List[float] = [img_col_width_char] + [0] * (len(COLUMNS) - 1)
        avg_char_w: float = font.getlength('A')

        for idx, col in enumerate(COLUMNS):
            if idx == img_idx:
                continue

            header_w_px, _ = get_text_px(col.header, font)
            char_w: float = header_w_px / avg_char_w + 0.1
            col_widths[idx] = max(col_widths[idx], char_w)

            for book in group_books:
                val: Any = col.getter(book)
                text: str = f'{int(val):,}원' if col.fmt_key == 'price' else str(val or '')

                w_px, _ = get_text_px(text, font)
                char_w = w_px / avg_char_w + 0.1

                col_widths[idx] = max(col_widths[idx], char_w)

            col_widths[idx] = min(col_widths[idx], 60)

        for idx, width in enumerate(col_widths):
            worksheet.set_column(idx, idx, width)
            worksheet.write(0, idx, COLUMNS[idx].header, fm.get('header'))

        row_heights: List[float] = [font_size_pt * 1.7]

        for book in group_books:
            max_h: float = font_size_pt * 1.7

            for idx, col in enumerate(COLUMNS):
                if idx == img_idx:
                    cell_h: float = 112
                else:
                    lines: int = (str(col.getter(book)) or '').count('\n') + 1
                    cell_h = lines * font_size_pt * 1.7

                max_h = max(max_h, cell_h)

            row_heights.append(max_h)

        for r, h in enumerate(row_heights):
            worksheet.set_row(r, h)

        for r, book in enumerate(group_books, start=1):
            for idx, col in enumerate(COLUMNS):
                if idx == img_idx:
                    continue

                val: Any = col.getter(book)

                if idx == 11:
                    if book.library_status == LibraryBookStatus.EXISTS and book.key and book.species_key:
                        url = f"https://read365.edunet.net/PureScreen/SearchDetail?bookKey={book.key}&speciesKey={book.species_key}&provCode={prov_code}&neisCode={neis_code}&schoolName={school_name}&fromSchool=true"
                        worksheet.write_url(r, idx, url, fm.get('hyperlink'), string='링크')
                    else:
                        worksheet.write(r, idx, val, fm.get(col.fmt_key or 'center'))
                elif idx == 1:
                    if book.item_id:
                        url = f"https://www.aladin.co.kr/shop/wproduct.aspx?ItemId={book.item_id}"
                        worksheet.write_url(r, idx, url, fm.get('hyperlink'), string=val)
                    else:
                        worksheet.write(r, idx, val, fm.get(col.fmt_key or 'center'))
                elif col.fmt_key == 'price':
                    worksheet.write_number(r, idx, val, fm.get('price'))
                elif col.fmt_key == 'sales_point':
                    worksheet.write_number(r, idx, val, fm.get('sales_point'))
                else:
                    worksheet.write(r, idx, val, fm.get(col.fmt_key or 'center'))

        for r, book in enumerate(group_books, start=1):
            if book.cover:
                cell_w_px: int = col_to_px(col_widths[img_idx])
                cell_h_px: int = row_to_px(row_heights[r])

                im: Image.Image = Image.open(book.cover)
                buf: BytesIO = BytesIO()

                im_resized: Image.Image = im.resize((cell_w_px, cell_h_px), Image.Resampling.LANCZOS)
                im_resized.save(buf, format='PNG')

                buf.seek(0)

                worksheet.insert_image(r, img_idx, f"{book.isbn13}.png", {'image_data': buf, 'x_offset': 0, 'y_offset': 0, 'positioning': 1})
        
        # 도서관 소장 도서는 노란색으로 행 강조
        worksheet.conditional_format(1, 0, len(group_books), len(COLUMNS)-1, {
            'type': 'formula',
            'criteria': '=$L2="링크"',
            'format': fm.get('highlight')
        })

        # 5년 이전 출판 도서는 연두색으로 행 강조
        current_year = datetime.now().year
        worksheet.conditional_format(1, 0, len(group_books), len(COLUMNS)-1, {
            'type': 'formula',
            'criteria': f'=AND(ISNUMBER(VALUE(LEFT($G2,4))), VALUE(LEFT($G2,4))<>{current_year}, VALUE(LEFT($G2,4))<{current_year - 4})',
            'format': fm.get('oldbook')
        })

    workbook.close()
    print(f"@@@@@ 엑셀 파일({output})을 저장했어요.")

if __name__ == "__main__":
    if not os.path.exists("config.yml"):
        with open("config.yml", "w", encoding="utf-8") as f:
            yaml.dump({
                "aladinKey": "write here",
                "libraryLink": "write here",
                "outputFileName": "output.xlsx"
            }, f, allow_unicode=True)

        print(f"config.yml 파일이 존재하지 않아 기본 템플릿을 생성했어요. 값을 채워넣고 다시 실행해주세요.")
        sys.exit(1)

    with open("config.yml", encoding="utf-8") as f:
        config: Dict[str, Any] = yaml.safe_load(f)
        required_keys: List[str] = ["aladinKey", "libraryLink", "outputFileName"]
        for k in required_keys:
            if k not in config or not config[k]:
                print(f"콘피그 파일에 '{k}' 값이 비어 있어요.")
                sys.exit(1)

    query_params: Dict[str, List[str]] = parse_qs(urlparse(config["libraryLink"]).query)

    SCHOOL_NAME: Optional[str] = query_params.get('schoolName', [None])[0]
    PROV_CODE: Optional[str] = query_params.get('provCode', [None])[0]
    NEIS_CODE: Optional[str] = query_params.get('neisCode', [None])[0]

    if SCHOOL_NAME is None or PROV_CODE is None or NEIS_CODE is None:
        print("올바르지 않은 도서관 링크가 제공되었어요. 프로그램의 설명을 참고하여 올바른 링크를 입력해주세요.")
        sys.exit(1)

    print(f"@@@@@ 학교 정보를 로딩했어요: {SCHOOL_NAME} (교육청 코드: {PROV_CODE}, 나이스 코드: {NEIS_CODE})")

    ALADIN_API_KEY: str = config["aladinKey"]
    OUTPUT_XLSX_FILE: str = config["outputFileName"]

    DEFAULT_FONT_SIZE_PT: int = 11
    DESCRIPTION_WRAP_WIDTH: int = 25

    INPUT_XLSX_FILE: str = "list.xlsx"

    books_to_check: List[Book] = []

    if not os.path.exists(INPUT_XLSX_FILE):
        workbook: openpyxl.Workbook = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        sheet['A1'] = 'ISBN13'
        sheet['B1'] = '시트'
        sheet['C1'] = '메모 (선택)'

        workbook.save(INPUT_XLSX_FILE)

        print(f"'{INPUT_XLSX_FILE}' 파일이 존재하지 않아 생성했어요. A열에 ISBN13, B열에 시트, C열에 메모를 입력해주세요.")
        sys.exit(1)

    try:
        workbook: openpyxl.Workbook = openpyxl.load_workbook(INPUT_XLSX_FILE, data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=3), start=1):
            isbn_cell: openpyxl.cell.cell.Cell = row[0]
            sheet_cell: Optional[openpyxl.cell.cell.Cell] = row[1] if len(row) > 1 else None
            memo_cell: Optional[openpyxl.cell.cell.Cell] = row[2] if len(row) > 2 else None

            if isbn_cell.value:
                raw = str(isbn_cell.value).strip()
                sheet_name = str(sheet_cell.value).strip() if sheet_cell and sheet_cell.value else ''
                memo = str(memo_cell.value).strip() if memo_cell and memo_cell.value else ''

                if not sheet_name:
                    print(f"'{raw}' 책이 시트가 지정되지 않았어요.")
                    sys.exit(1)

                if raw.startswith("http"):
                    parsed = urlparse(raw)
                    query = parse_qs(parsed.query)
                    id_list = query.get("ItemId") or query.get("itemId")
                    if not id_list or not id_list[0].isdigit():
                        print(f"'{raw}' 에서 ItemId를 추출할 수 없어요.")
                        sys.exit(1)
                    item_id = int(id_list[0])
                    book_isbn = ""
                else:
                    item_id = None
                    book_isbn = raw

                books_to_check.append(
                    Book(item_id=item_id, isbn13=book_isbn, sheet_name=sheet_name, memo=memo, order=row_idx)
                )

    except Exception as e:
        print(f"@@@@@ '{INPUT_XLSX_FILE}' 파일을 읽는 중 오류가 발생했어요: {e}")
        sys.exit(1)

    if not books_to_check:
        print("처리할 책 데이터가 없어요. 프로그램을 종료합니다.")
        sys.exit(1)

    create(books_to_check, OUTPUT_XLSX_FILE, DEFAULT_FONT_SIZE_PT, ALADIN_API_KEY, NEIS_CODE, PROV_CODE, SCHOOL_NAME)